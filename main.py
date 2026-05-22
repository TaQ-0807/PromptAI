import base64
import json
import random
import secrets
import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
from anthropic import Anthropic
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
import os

load_dotenv()

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

client = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

BASE_DIR      = Path(__file__).parent
CONV_DIR      = BASE_DIR / "conversations"
CODES_FILE    = BASE_DIR / "codes.json"
SESSIONS_FILE = BASE_DIR / "sessions.json"
try:
    CONV_DIR.mkdir(exist_ok=True)
except OSError:
    pass  # Vercel 等の読み取り専用環境では無視（Supabase を使用）

OWNER_USERNAME      = os.getenv("OWNER_USERNAME", "admin")
OWNER_PASSWORD      = os.getenv("OWNER_PASSWORD", "")
SAFE_CHARS          = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
SESSION_EXPIRE_DAYS = 30

# ─── Supabase（オプション：環境変数が設定された場合のみ有効）────────
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
supa = None

if SUPABASE_URL and SUPABASE_KEY:
    try:
        from supabase import create_client
        supa = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("✓ Supabase に接続しました")
    except Exception as e:
        print(f"Supabase 接続エラー: {e} → ファイルストレージにフォールバック")


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _parse_dt(s: str) -> datetime:
    """ISO datetime 文字列をタイムゾーン付きで解析"""
    return datetime.fromisoformat(s.replace("Z", "+00:00"))


# ─── ファイルベース ストレージ（ローカル開発用）────────────────────

def _load_sessions_file() -> dict:
    if not SESSIONS_FILE.exists():
        return {}
    try:
        return json.loads(SESSIONS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_sessions_file(data: dict):
    SESSIONS_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _load_codes_file() -> dict:
    if not CODES_FILE.exists():
        return {}
    try:
        return json.loads(CODES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_codes_file(data: dict):
    CODES_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ─── セッション管理 ────────────────────────────────────────────────

def create_session() -> str:
    token = secrets.token_hex(32)
    now = _now_utc()
    expires_at = (now + timedelta(days=SESSION_EXPIRE_DAYS)).isoformat()

    if supa:
        # 期限切れセッションを削除
        supa.table("sessions").delete().lt("expires_at", now.isoformat()).execute()
        supa.table("sessions").insert({
            "token": token,
            "expires_at": expires_at,
        }).execute()
    else:
        sessions = _load_sessions_file()
        sessions = {k: v for k, v in sessions.items()
                    if datetime.fromisoformat(v["expires_at"]) > datetime.now()}
        sessions[token] = {
            "created_at": now.isoformat(),
            "expires_at": expires_at,
        }
        _save_sessions_file(sessions)
    return token


def validate_session(token: str) -> bool:
    if not token:
        return False
    if supa:
        r = supa.table("sessions").select("expires_at").eq("token", token).execute()
        if not r.data:
            return False
        return _now_utc() < _parse_dt(r.data[0]["expires_at"])
    else:
        sessions = _load_sessions_file()
        if token not in sessions:
            return False
        return datetime.now() < datetime.fromisoformat(sessions[token]["expires_at"])


def invalidate_session(token: str):
    if supa:
        supa.table("sessions").delete().eq("token", token).execute()
    else:
        sessions = _load_sessions_file()
        if token in sessions:
            del sessions[token]
            _save_sessions_file(sessions)


# ─── コード管理 ────────────────────────────────────────────────────

def load_codes() -> dict:
    if supa:
        r = supa.table("codes").select("*").execute()
        return {
            row["code"]: {
                "used":       row["used"],
                "created_at": row["created_at"],
                "used_at":    row["used_at"],
            }
            for row in r.data
        }
    else:
        return _load_codes_file()


def auth_required() -> bool:
    """認証が必要かどうかを判定"""
    if supa:
        return bool(OWNER_PASSWORD)
    return bool(OWNER_PASSWORD) or CODES_FILE.exists()


def make_code() -> str:
    return "".join(random.choices(SAFE_CHARS, k=8))


SYSTEM_PROMPT = """あなたはPromptoAIのAIコンシェルジュ「Prompto」です。
クライアントが作りたいツールの要件を徹底的にヒアリングし、Claude Codeが一発で高い完成度のコードを生成できる、精密なプロンプトを作るのがあなたの使命です。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【最重要原則】機能の完全性を最優先する
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
「元のツールを参考にWebアプリ化してほしい」という依頼では、
元ツールの機能が削られたり改変されることが最大のリスクです。
あなたはこのリスクを徹底的につぶすことを最優先にしてください。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【STEP 0】ファイルが添付された場合（最優先で実行）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ExcelやPDF・画像などのファイルが添付されたら、まず以下を行ってください：

■ 徹底的な内容分析
添付ファイルを精密に読み込み、以下を全て洗い出してください：

【Excelの場合】
- シート構成（シート名・役割・データ構造）
- 全ての入力項目（セル・ドロップダウン・入力規則）
- 全ての計算式・自動計算ロジック
- 条件付き書式・色分けのルール
- ボタン・マクロ・自動化処理
- データの参照関係（シート間連携）
- 出力・印刷・エクスポート機能
- ユーザーの操作フロー（どの順番で何を入力するか）

【画像・UIの場合】
- 表示されている全ての要素・ボタン・フォーム
- レイアウト・配置の意図
- 操作の流れ

■ 分析結果の報告と確認
分析した内容をクライアントに提示し、必ず次の確認をしてください：
「以下の機能・要素を確認しました。抜け漏れや補足はありますか？」

■ 隠れた機能の確認
「他に使っているシートや機能、マクロ、外部連携などはありますか？」と必ず確認する。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【STEP 1】最初の質問（ファイルなしの場合）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
温かく迎え、「どんなツールを作りたいですか？」とだけ聞いてください。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【STEP 2】要件ヒアリング（深掘り）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1回に1〜2個の質問で、以下を引き出してください：

■ 機能の優先順位（最重要）
「絶対に必要な機能」と「あれば嬉しい機能」を明確に分ける。
特に既存ツールのWebアプリ化では：
- 「元のツールで絶対に変えたくない機能・操作感はどれですか？」
- 「逆に、この機会に改善したい部分はありますか？」
- 「今回は実装しなくて良いと割り切れる機能はありますか？」

■ ユーザーと操作フロー
- 誰が、どの場面で使うか
- 1回の操作でどんな流れをたどるか（ステップ順に）
- 複数人で使う場合、役割の違いはあるか

■ データと連携
- 入力するデータの種類・量
- データの保存先（ローカル / クラウド / DB）
- 既存システムとの連携の有無
- エクスポート・印刷・共有の必要性

■ 技術・環境
- 使用者のITリテラシー
- 動作環境（PC / スマホ / タブレット）
- インターネット接続の有無
- 技術スタックの希望（なければ最適なものを提案）

■ デザイン・UI
- 既存ツールのどのUIを引き継ぎたいか
- 追加したいUI改善点

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【STEP 3】機能リストの最終確認（プロンプト生成前に必須）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
プロンプトを生成する前に、必ず以下を行ってください：

洗い出した全機能をリスト形式でクライアントに提示し、
「この全機能を実装する方向で進めてよいですか？優先度の変更はありますか？」
と確認してから、承諾を得てプロンプト生成に進んでください。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【STEP 4】プロンプト生成
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
承諾を得たら、以下のフォーマットで出力してください。
必ず <prompt> タグと </prompt> タグで囲んでください。

<prompt>
# [プロジェクト名] 開発仕様

## 概要
[ツールの目的・背景・誰が何のために使うかを3〜5文で詳しく]

## 対象ユーザー
[具体的なユーザー像・ITリテラシー・利用場面]

## 技術スタック
- フロントエンド：[...]
- バックエンド：[...]
- データベース：[...]
- その他ライブラリ：[...]

## ディレクトリ構成
```
project/
├── ...
```

## 機能要件

### 🔴 必須機能（全て実装すること・省略・簡略化禁止）
以下の機能は元のツールから引き継ぐ中核機能です。一つも省略せず、完全に実装してください。
1. **[機能名]**：[操作方法・計算ロジック・表示形式まで詳細に記述]
2. **[機能名]**：[同上]
※ 機能数に応じてリストを展開する

### 🟡 改善機能（元ツールから改善・モダン化する点）
1. **[機能名]**：[元の動作と、どう改善するかを明記]

### 🟢 追加機能（余裕があれば実装）
1. **[機能名]**：[詳細]

## 画面・UI仕様

### 画面一覧
[全画面・全タブ・全モーダルを列挙]

### 各画面の詳細
#### [画面名]
- 表示する情報：[...]
- 操作できる内容：[...]
- 遷移先：[...]

## 操作フロー（ユーザーが実際に行う手順）
1. [ステップ1]
2. [ステップ2]
3. [以降続ける]

## データモデル
[全テーブル・全フィールド・型・制約を詳細に記述]

## 計算・自動処理ロジック
[全ての計算式・自動入力・条件分岐・バリデーションを詳細に記述]

## 外部連携・エクスポート
[ファイル出力・印刷・API連携などの仕様]

## 実装の注意点
- [注意点1：特に元ツールから変えてはいけない挙動を明記]
- [注意点2]

## 今回のスコープ外（意図的に省く機能）
- [除外する機能と、その理由]

---
## 実装指示

上記仕様に従って実装を開始してください。

【厳守事項】
- 「🔴 必須機能」は一つも省略・簡略化してはいけません
- 実装が複雑でも、分割して全て実装してください
- 分からない仕様があれば省略せず、コメントで TODO として残してください
- 実装完了後、必須機能チェックリストを出力して全機能の実装を確認してください

実装順序：
1. プロジェクト構成・環境構築
2. データモデル・DB設計
3. 🔴 必須機能（コア機能から順に）
4. 画面・UI実装
5. 🟡 改善機能
6. テスト・動作確認
</prompt>

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【会話全体を通じた重要ルール】
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- 必ず日本語で話す
- 質問は一度に1〜2個まで（情報過多にしない）
- 専門用語はわかりやすく説明する
- クライアントが答えに詰まったら具体例を提示して助ける
- フレンドリーかつプロフェッショナルなトーンを保つ
- <prompt>タグはプロンプト生成時のみ使用する
- プロンプト生成後も修正依頼に応じてプロンプトを更新できる
- 「だいたいこんな感じで」という曖昧な回答には、具体的な選択肢を示して確認する"""


# ─── Helpers ───────────────────────────────────────────────────────

def process_excel(file_bytes: bytes, filename: str) -> str:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    lines = [f"【Excelファイル: {filename}】"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== シート: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                lines.append("\t".join("" if v is None else str(v) for v in row))
    return "\n".join(lines)


def get_image_media_type(content_type: str, filename: str) -> str:
    mapping = {
        "image/jpeg": "image/jpeg",
        "image/jpg":  "image/jpeg",
        "image/png":  "image/png",
        "image/gif":  "image/gif",
        "image/webp": "image/webp",
    }
    ct = content_type.lower()
    if ct in mapping:
        return mapping[ct]
    ext = filename.lower().rsplit(".", 1)[-1]
    return {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
            "gif": "image/gif", "webp": "image/webp"}.get(ext, "image/jpeg")


# ─── Models ────────────────────────────────────────────────────────

class SaveRequest(BaseModel):
    id: Optional[str] = None
    messages: list[dict]


# ─── Routes ────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


# ─── Auth ──────────────────────────────────────────────────────────

@app.post("/auth/login")
async def auth_login(body: dict):
    """オーナーログイン"""
    if not OWNER_PASSWORD:
        return JSONResponse(status_code=503, content={"error": "no_owner_configured"})
    username = body.get("username", "").strip()
    password = body.get("password", "")
    if username == OWNER_USERNAME and password == OWNER_PASSWORD:
        token = create_session()
        return {"ok": True, "token": token}
    return JSONResponse(status_code=401, content={"error": "invalid_credentials"})


@app.post("/auth/logout")
async def auth_logout(body: dict):
    """オーナーログアウト"""
    token = body.get("token", "")
    invalidate_session(token)
    return {"ok": True}


@app.post("/auth/verify")
async def auth_verify(body: dict):
    """セッション or 招待コードの検証"""
    # オーナーセッション確認
    token = body.get("token", "")
    if validate_session(token):
        return {"ok": True, "owner_mode": True}
    # 認証不要モード
    if not auth_required():
        return {"ok": True, "owner_mode": True}
    # 招待コード確認
    code = body.get("code", "").strip().upper()
    if not code:
        return JSONResponse(status_code=401, content={"error": "no_code"})

    if supa:
        r = supa.table("codes").select("used").eq("code", code).execute()
        if not r.data:
            return JSONResponse(status_code=401, content={"error": "invalid"})
        if r.data[0]["used"]:
            return JSONResponse(status_code=403, content={"error": "used"})
    else:
        codes = _load_codes_file()
        if code not in codes:
            return JSONResponse(status_code=401, content={"error": "invalid"})
        if codes[code]["used"]:
            return JSONResponse(status_code=403, content={"error": "used"})

    return {"ok": True}


# ─── Admin ─────────────────────────────────────────────────────────

@app.get("/admin/codes")
async def admin_list_codes(request: Request):
    token = request.headers.get("X-Owner-Token", "")
    if not validate_session(token):
        return JSONResponse(status_code=401, content={"error": "unauthorized"})
    return load_codes()


@app.post("/admin/codes/generate")
async def admin_generate_codes(body: dict, request: Request):
    token = request.headers.get("X-Owner-Token", "")
    if not validate_session(token):
        return JSONResponse(status_code=401, content={"error": "unauthorized"})
    n = max(1, min(int(body.get("count", 1)), 50))
    codes = load_codes()
    new_codes = []
    now = _now_utc().isoformat()
    rows_to_insert = []

    for _ in range(n):
        code = make_code()
        while code in codes:
            code = make_code()
        codes[code] = {"used": False, "created_at": now, "used_at": None}
        new_codes.append(code)
        rows_to_insert.append({"code": code, "used": False, "used_at": None})

    if supa:
        supa.table("codes").insert(rows_to_insert).execute()
    else:
        _save_codes_file(codes)

    return {"ok": True, "codes": new_codes}


@app.delete("/admin/codes/{code}")
async def admin_delete_code(code: str, request: Request):
    token = request.headers.get("X-Owner-Token", "")
    if not validate_session(token):
        return JSONResponse(status_code=401, content={"error": "unauthorized"})
    code_upper = code.upper()

    if supa:
        supa.table("codes").delete().eq("code", code_upper).execute()
    else:
        codes = _load_codes_file()
        if code_upper in codes:
            del codes[code_upper]
            _save_codes_file(codes)

    return {"ok": True}


# ─── Chat ──────────────────────────────────────────────────────────

@app.post("/chat")
async def chat(
    messages: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    code: str = Form(default=""),
    session_token: str = Form(default=""),
):
    is_owner = validate_session(session_token)
    c = code.strip().upper()

    # ゲストのみコード認証（オーナーはスキップ）
    if not is_owner and auth_required():
        if supa:
            r = supa.table("codes").select("used").eq("code", c).execute()
            if not r.data:
                return JSONResponse(status_code=401, content={"error": "invalid_code"})
            if r.data[0]["used"]:
                return JSONResponse(status_code=403, content={"error": "code_used"})
        else:
            codes = _load_codes_file()
            if c not in codes:
                return JSONResponse(status_code=401, content={"error": "invalid_code"})
            if codes[c]["used"]:
                return JSONResponse(status_code=403, content={"error": "code_used"})

    msgs = json.loads(messages)

    if files:
        file_blocks = []
        for file in files:
            file_bytes = await file.read()
            ct   = (file.content_type or "").lower()
            name = file.filename or ""

            if ct.startswith("image/") or name.lower().rsplit(".", 1)[-1] in ("jpg", "jpeg", "png", "gif", "webp"):
                media_type = get_image_media_type(ct, name)
                file_blocks.append({
                    "type": "image",
                    "source": {
                        "type":       "base64",
                        "media_type": media_type,
                        "data":       base64.standard_b64encode(file_bytes).decode(),
                    },
                })
            elif ct == "application/pdf" or name.lower().endswith(".pdf"):
                file_blocks.append({
                    "type": "document",
                    "source": {
                        "type":       "base64",
                        "media_type": "application/pdf",
                        "data":       base64.standard_b64encode(file_bytes).decode(),
                    },
                })
            elif "spreadsheet" in ct or "excel" in ct or name.lower().endswith((".xlsx", ".xls", ".xlsm")):
                file_blocks.append({
                    "type": "text",
                    "text": process_excel(file_bytes, name),
                })

        last = msgs[-1]
        text_content = last["content"] if isinstance(last["content"], str) else ""
        content = []
        if text_content:
            content.append({"type": "text", "text": text_content})
        content.extend(file_blocks)
        msgs[-1] = {"role": "user", "content": content}

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        messages=msgs,
    )
    response_text = response.content[0].text

    # ゲストのみコード消費（オーナーは無制限）
    code_consumed = False
    if not is_owner and auth_required() and "<prompt>" in response_text:
        if supa:
            r = supa.table("codes").select("used").eq("code", c).execute()
            if r.data and not r.data[0]["used"]:
                supa.table("codes").update({
                    "used":    True,
                    "used_at": _now_utc().isoformat(),
                }).eq("code", c).execute()
                code_consumed = True
        else:
            codes = _load_codes_file()
            if c in codes and not codes[c]["used"]:
                codes[c]["used"]    = True
                codes[c]["used_at"] = datetime.now().isoformat()
                _save_codes_file(codes)
                code_consumed = True

    return {"content": response_text, "code_consumed": code_consumed}


# ─── Conversation history ───────────────────────────────────────────

@app.get("/conversations")
async def list_conversations():
    if supa:
        r = supa.table("conversations").select(
            "id, title, created_at, updated_at"
        ).order("updated_at", desc=True).execute()
        return r.data
    else:
        convs = []
        files = sorted(CONV_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
        for f in files:
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                convs.append({
                    "id":         data["id"],
                    "title":      data["title"],
                    "created_at": data["created_at"],
                    "updated_at": data["updated_at"],
                })
            except Exception:
                continue
        return convs


@app.post("/conversations/save")
async def save_conversation(body: SaveRequest):
    conv_id = body.id or str(uuid.uuid4())

    first_msg = next((m["content"] for m in body.messages if m["role"] == "user"), "無題の会話")
    if isinstance(first_msg, list):
        first_msg = next((b["text"] for b in first_msg if b.get("type") == "text"), "無題の会話")
    title = (first_msg[:38] + "…") if len(first_msg) > 38 else first_msg

    now = _now_utc().isoformat()

    if supa:
        # 既存レコードから created_at と title を維持
        r = supa.table("conversations").select("created_at, title").eq("id", conv_id).execute()
        created_at = r.data[0]["created_at"] if r.data else now
        if r.data:
            title = r.data[0]["title"]

        supa.table("conversations").upsert({
            "id":         conv_id,
            "title":      title,
            "messages":   body.messages,
            "created_at": created_at,
            "updated_at": now,
        }).execute()
    else:
        existing_path = CONV_DIR / f"{conv_id}.json"
        created_at = now
        if existing_path.exists():
            try:
                existing   = json.loads(existing_path.read_text(encoding="utf-8"))
                created_at = existing.get("created_at", now)
                title      = existing.get("title", title)
            except Exception:
                pass

        data = {
            "id":         conv_id,
            "title":      title,
            "created_at": created_at,
            "updated_at": now,
            "messages":   body.messages,
        }
        existing_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"id": conv_id, "title": title}


@app.get("/conversations/{conv_id}")
async def get_conversation(conv_id: str):
    if supa:
        r = supa.table("conversations").select("*").eq("id", conv_id).execute()
        if not r.data:
            return JSONResponse(status_code=404, content={"error": "not found"})
        return r.data[0]
    else:
        path = CONV_DIR / f"{conv_id}.json"
        if not path.exists():
            return JSONResponse(status_code=404, content={"error": "not found"})
        return json.loads(path.read_text(encoding="utf-8"))


@app.delete("/conversations/{conv_id}")
async def delete_conversation(conv_id: str):
    if supa:
        supa.table("conversations").delete().eq("id", conv_id).execute()
    else:
        path = CONV_DIR / f"{conv_id}.json"
        if path.exists():
            path.unlink()
    return {"ok": True}
